import openpyxl as xl
wb = xl.load_workbook("Viva.xlsx")
sheet = wb['Sheet2']
print(sheet.max_row)
for row in sheet.values:
    for value in row:
        if (value == 8178240640):
            print(row)
print(sheet.dimensions)

for row in sheet.iter_rows(min_row=2, max_row=6):
    for cell in row:
        print(cell.value, end=" ")
    print()